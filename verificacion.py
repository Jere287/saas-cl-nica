"""
verificacion.py — Genera el Excel de VERIFICACIÓN en el formato EXACTO de la
hoja de cálculos manuales del área de calidad.

En vez de recrear el formato (fuentes, bordes, anchos y combinaciones del
equipo son irregulares y fáciles de errar), se usa una PLANTILLA
(plantillas/verificacion.xlsx) que es el archivo del equipo con sus fórmulas
intactas. La app solo escribe:
  - los valores de cabecera de la pieza (columna E, que las demás columnas
    propagan por fórmula), y
  - la columna Impedance con las mediciones crudas.
Las fórmulas de la plantilla (=AVERAGE, =STDEV, =MAX-MIN) reproducen el
promedio, la desviación estándar y el rango exactamente como el cálculo manual.

Requiere las mediciones crudas, que el parser conserva en
datos['pruebas'][prueba][i]['mediciones'].
"""
import os
import shutil

from openpyxl import load_workbook

PLANTILLA = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         'plantillas', 'verificacion.xlsx')

# Posiciones fijas del formato del equipo (nunca cambian).
OPT_BLOCKS = [2, 6, 11, 15, 20, 24, 29, 33, 38, 42]  # columna inicial de cada canal óptico
ELE_BLOCKS = [2, 7, 13, 18]                           # columna inicial de cada frecuencia
DATA_START = 15
N_OPT = 20   # 20 mediciones por canal óptico
N_ELE = 60   # 20 mediciones x 3 electrodos por frecuencia
# Celdas de cabecera de la pieza (columna E=5); las demás columnas las propagan.
HDR = {2: 'Test', 3: 'Batch', 4: 'Equipment number', 5: 'SN',
       7: 'MAC', 9: 'Test date', 10: 'Test responsible', 11: 'Test Standard'}


def _rellenar_cabecera(ws, cab):
    """Escribe en la columna E los datos de la pieza; la plantilla propaga el
    resto de bloques por fórmula (=IF(E..<>"",E..,""))."""
    valores = {2: cab.get('test', 'Current'), 3: cab.get('batch', ''),
               4: cab.get('equipo', ''), 5: cab.get('sn', ''),
               7: cab.get('mac', ''), 9: cab.get('fecha', ''),
               10: cab.get('responsable', ''), 11: cab.get('estandar', '')}
    for fila, val in valores.items():
        if val not in (None, ''):
            ws.cell(row=fila, column=5).value = val


def _rellenar_impedancias(ws, bloques, offset, series):
    """series: lista (una por bloque) de listas de mediciones crudas.
    offset: columna Impedance dentro del bloque (óptico +2, eléctrico +3)."""
    for bcol, valores in zip(bloques, series):
        for i, v in enumerate(valores):
            ws.cell(row=DATA_START + i, column=bcol + offset).value = float(v)


def generar_verificacion(ruta_salida, pieza, cabecera):
    """pieza: dict con las mediciones crudas ya ordenadas:
        {'optico': [ [20 valores] x10 canales ],
         'electrico': [ [60 valores] x4 frecuencias ]}
    cabecera: dict con sn, batch, mac, fecha, responsable, estandar, equipo, test.
    Devuelve la ruta guardada. Las fórmulas calculan promedio/desv/rango al abrir."""
    if not os.path.exists(PLANTILLA):
        raise FileNotFoundError(f'No se encontró la plantilla de verificación: {PLANTILLA}')
    # se copia la plantilla y se rellena (así nunca se altera el original)
    shutil.copyfile(PLANTILLA, ruta_salida)
    wb = load_workbook(ruta_salida)

    wo = wb['Optico']
    _rellenar_cabecera(wo, cabecera)
    _rellenar_impedancias(wo, OPT_BLOCKS, 2, pieza['optico'])

    we = wb['Electrico']
    _rellenar_cabecera(we, cabecera)
    _rellenar_impedancias(we, ELE_BLOCKS, 3, pieza['electrico'])

    wb.save(ruta_salida)
    return ruta_salida
