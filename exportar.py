"""
exportar.py — Exporta los datos finales de un lote a un archivo Excel (.xlsx)
como documento formal de liberación de lote, en estilo sobrio de documento
regulatorio: blanco y negro, gris claro solo para estructura (encabezados y
etiquetas), sin colores decorativos. El estado se lee por el TEXTO
(PASÓ / NO PASÓ / SIN ESTÁNDAR / RECHAZADO), no por colores.

Tres hojas:
  1. Reporte de liberación : membrete, datos del lote, DICTAMEN, resumen de
                             resultados, criterios de evaluación y aprobaciones.
  2. Resultados por pieza  : tabla agrupada por veredicto (primero las que
                             requieren atención).
  3. Verificación de cálculos : respaldo crudo por canal (media, desviación,
                             rango y estado).

Las tres hojas van sin cuadrícula y salen listas para imprimir (ajuste a lo
ancho y "Página X de Y" en el pie). El nombre de la empresa se personaliza en
reporte.py (EMPRESA / SUBTITULO).
"""
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from reporte import EMPRESA, SUBTITULO

# ---- estilo documental: negro + dos grises ----
GRIS_NOTA = '595959'      # notas y textos secundarios
FILL_ETIQUETA = PatternFill('solid', fgColor='F2F2F2')   # etiquetas y encabezados
FILL_GRUPO = PatternFill('solid', fgColor='D9D9D9')      # filas de grupo/sección

ROTULOS = {
    'NO PASO':    'NO PASARON — repetir la prueba',
    'INCOMPLETO': 'SIN ESTÁNDAR — completar el perfil',
    'RECHAZADO':  'RECHAZADOS — archivo inválido',
    'PASO':       'PASARON',
}
ORDEN_GRUPOS = ['NO PASO', 'INCOMPLETO', 'RECHAZADO', 'PASO']
VER_TXT = {'PASO': 'PASÓ', 'NO PASO': 'NO PASÓ',
           'INCOMPLETO': 'SIN ESTÁNDAR', 'RECHAZADO': 'RECHAZADO'}

MINI = Font(name='Arial', size=8.5, color=GRIS_NOTA)
HEAD = Font(name='Arial', bold=True, size=10)
BOLD = Font(name='Arial', bold=True, size=10)
NORM = Font(name='Arial', size=10)
GRUPO_F = Font(name='Arial', bold=True, size=10)
SECCION = Font(name='Arial', bold=True, size=11)
CENTER = Alignment(horizontal='center', vertical='center')
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
IZQ = Alignment(horizontal='left', vertical='center')
IZQ_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
DER = Alignment(horizontal='right', vertical='center')
thin = Side(style='thin', color='A6A6A6')
negro = Side(style='medium', color='000000')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
LINEA_SECCION = Border(bottom=Side(style='thin', color='000000'))
LINEA_MEMBRETE = Border(bottom=negro)


# =================== utilidades ===================
def _estado_prueba(detalle, prueba):
    fs = detalle.get(prueba, [])
    if not fs:
        return '—'
    if any(x['resultado'] == 'FALLA' for x in fs):
        return 'No pasó'
    # sin ninguna falla real: si algun canal no tiene estandar, no se puede
    # afirmar que paso -> se avisa que falta definir el limite (no es una falla)
    if any(x['resultado'] == 'SIN ESTANDAR' for x in fs):
        return 'Sin estándar'
    return 'Pasó'


def _estado(celda, texto):
    """Estado en texto plano: negrita solo cuando requiere atención."""
    celda.alignment = CENTER
    celda.border = BORDER
    if texto in ('No pasó', 'NO PASÓ', 'FALLA', 'Falla', 'RECHAZADO', 'Rechazado'):
        celda.font = BOLD
    else:
        celda.font = NORM


def _caja(ws, fila, c1, c2, fill=None):
    """Borde (y fondo opcional) a todas las celdas de un rango de una fila;
    con celdas combinadas el borde debe ponerse celda por celda."""
    for c in range(c1, c2 + 1):
        celda = ws.cell(row=fila, column=c)
        celda.border = BORDER
        if fill is not None:
            celda.fill = fill


def _merge(ws, fila, c1, c2, valor=None, font=None, align=None, fill=None, borde=True):
    ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
    celda = ws.cell(row=fila, column=c1)
    if valor is not None:
        celda.value = valor
    if font:
        celda.font = font
    if align:
        celda.alignment = align
    for c in range(c1, c2 + 1):
        cc = ws.cell(row=fila, column=c)
        if fill is not None:
            cc.fill = fill
        if borde:
            cc.border = BORDER
    return celda


def _marco(ws, r1, r2, c1, c2):
    """Marco exterior en línea negra media para un bloque (p. ej. el dictamen)."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            izq = negro if c == c1 else thin
            der = negro if c == c2 else thin
            arr = negro if r == r1 else thin
            aba = negro if r == r2 else thin
            ws.cell(row=r, column=c).border = Border(left=izq, right=der, top=arr, bottom=aba)


def _grupo(ws, fila, c1, c2, texto):
    _merge(ws, fila, c1, c2, texto, font=GRUPO_F, align=IZQ, fill=FILL_GRUPO)
    ws.row_dimensions[fila].height = 16


def _membrete(ws, ncols, titulo_hoja, batch, codigo):
    """Membrete documental: empresa y título de la hoja, con línea negra debajo."""
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 13
    mitad = ncols // 2
    _merge(ws, 2, 1, mitad, EMPRESA,
           font=Font(name='Arial', bold=True, size=13), align=IZQ, borde=False)
    _merge(ws, 2, mitad + 1, ncols, titulo_hoja,
           font=Font(name='Arial', bold=True, size=11), align=DER, borde=False)
    _merge(ws, 3, 1, mitad, SUBTITULO, font=MINI, align=IZQ, borde=False)
    _merge(ws, 3, mitad + 1, ncols,
           ' · '.join(p for p in [f'Documento {codigo}', batch.get('nombre', ''), 'CONFIDENCIAL'] if p),
           font=MINI, align=DER, borde=False)
    for c in range(1, ncols + 1):
        ws.cell(row=3, column=c).border = LINEA_MEMBRETE


def _titulo_seccion(ws, fila, c1, c2, texto):
    ws.merge_cells(start_row=fila, start_column=c1, end_row=fila, end_column=c2)
    t = ws.cell(row=fila, column=c1, value=texto)
    t.font = SECCION; t.alignment = IZQ
    for c in range(c1, c2 + 1):
        ws.cell(row=fila, column=c).border = LINEA_SECCION
    ws.row_dimensions[fila].height = 18


def _campo(ws, fila, col, etiqueta, valor, val_hasta, bold=False):
    e = ws.cell(row=fila, column=col, value=etiqueta)
    e.font = BOLD; e.fill = FILL_ETIQUETA; e.border = BORDER; e.alignment = IZQ
    return _merge(ws, fila, col + 1, val_hasta, valor,
                  font=Font(name='Arial', size=10, bold=bold), align=IZQ)


def _encabezado_tabla(ws, fila, celdas):
    """celdas: lista de (col_inicio, col_fin, texto)."""
    for c1, c2, texto in celdas:
        if c1 == c2:
            c = ws.cell(row=fila, column=c1, value=texto)
            c.font = HEAD; c.fill = FILL_ETIQUETA; c.alignment = CENTER; c.border = BORDER
        else:
            _merge(ws, fila, c1, c2, texto, font=HEAD, align=CENTER, fill=FILL_ETIQUETA)
    ws.row_dimensions[fila].height = 16


def _preparar_impresion(ws, orientacion='portrait', ncols=8):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = orientacion
    ws.page_setup.paperSize = 9  # carta/A4
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.left.text = f'{EMPRESA} · Sistema de Control de Calidad'
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = '808080'
    ws.oddFooter.right.text = 'Página &P de &N'
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = '808080'
    ws.print_area = f'A1:{get_column_letter(ncols)}{ws.max_row + 1}'


# =================== hoja 1: reporte de liberación ===================
def _hoja_portada(ws, batch, conteos, codigo):
    # columnas: A margen | B..G contenido | H margen
    for col, w in zip('ABCDEFGH', [2.5, 21, 19, 21, 19, 13, 13, 2.5]):
        ws.column_dimensions[col].width = w
    total, pasaron, fallaron, incompletos, rechazados = conteos
    tasa = round(100 * pasaron / total) if total else None

    _membrete(ws, 8, 'REPORTE DE LIBERACIÓN DE LOTE', batch, codigo)

    # ---- I. datos del lote ----
    _titulo_seccion(ws, 5, 2, 7, 'I. DATOS DEL LOTE')
    emision = datetime.now().strftime('%d/%m/%Y %H:%M')
    _campo(ws, 7, 2, 'Lote / corrida', batch.get('nombre', ''), 3, bold=True)
    _campo(ws, 7, 4, 'Fecha de evaluación', (batch.get('fecha', '') or '')[:16].replace('T', ' '), 5)
    _campo(ws, 8, 2, 'Perfil de estándar', batch.get('perfil', ''), 3)
    _campo(ws, 8, 4, 'Operador', batch.get('operador', ''), 5)
    _campo(ws, 9, 2, 'Piezas evaluadas', total, 3)
    _campo(ws, 9, 4, 'Emisión del reporte', emision, 5)
    _campo(ws, 10, 2, 'Método de evaluación',
           'Dos fases por canal: desviación estándar y media contra límites del perfil', 5)
    for f in (7, 8, 9, 10):
        ws.row_dimensions[f].height = 16

    # ---- II. dictamen ----
    _titulo_seccion(ws, 12, 2, 7, 'II. DICTAMEN DE LIBERACIÓN')
    if fallaron or rechazados:
        dictamen = 'LOTE NO CONFORME'
        motivo = (f'{fallaron} pieza(s) no cumplieron los límites del perfil'
                  + (f' y {rechazados} archivo(s) fueron rechazados por formato inválido' if rechazados else '')
                  + '. Se requiere repetir la prueba en el siguiente desechable y documentar la disposición.')
    elif incompletos:
        dictamen = 'EVALUACIÓN INCOMPLETA — NO LIBERADO'
        motivo = (f'{incompletos} pieza(s) tienen canales sin límite definido en el perfil "'
                  f"{batch.get('perfil', '')}\". Complete el estándar de calidad y reevalúe antes de liberar.")
    elif total:
        dictamen = 'LOTE CONFORME — LIBERADO'
        motivo = (f'Las {total} piezas evaluadas cumplen los límites de desviación y media '
                  'del perfil en todos los canales (pruebas óptica y eléctrica).')
    else:
        dictamen = 'SIN PIEZAS EVALUADAS'
        motivo = 'El lote no contiene resultados.'
    ws.row_dimensions[14].height = 26
    _merge(ws, 14, 2, 7, dictamen,
           font=Font(name='Arial', bold=True, size=14), align=CENTER, borde=False)
    ws.row_dimensions[15].height = 26
    _merge(ws, 15, 2, 7, motivo,
           font=Font(name='Arial', size=9.5), align=CENTER_WRAP, borde=False)
    _marco(ws, 14, 15, 2, 7)

    # ---- III. resumen de resultados ----
    _titulo_seccion(ws, 17, 2, 7, 'III. RESUMEN DE RESULTADOS')
    hr = 19
    _encabezado_tabla(ws, hr, [(2, 2, 'Resultado'), (3, 3, 'Piezas'), (4, 4, '% del total')])
    filas_kpi = [('Pasaron', pasaron), ('No pasaron', fallaron),
                 ('Sin estándar (incompletas)', incompletos),
                 ('Rechazadas (archivo inválido)', rechazados)]
    r = hr + 1
    for nombre, n in filas_kpi:
        _caja(ws, r, 2, 4)
        ws.cell(row=r, column=2, value=nombre).font = NORM
        cn = ws.cell(row=r, column=3, value=n)
        cn.alignment = CENTER
        cn.font = BOLD if (n and nombre != 'Pasaron') else NORM
        cp = ws.cell(row=r, column=4, value=(f'{100 * n / total:.0f}%' if total else '—'))
        cp.alignment = CENTER; cp.font = NORM
        r += 1
    _caja(ws, r, 2, 4)
    ws.cell(row=r, column=2, value='Total de piezas evaluadas').font = BOLD
    ct = ws.cell(row=r, column=3, value=total); ct.font = BOLD; ct.alignment = CENTER
    ws.cell(row=r, column=4, value='100%' if total else '—').alignment = CENTER

    # índice de aprobación a la derecha, en un marco sobrio
    ws.merge_cells(start_row=hr, start_column=5, end_row=r - 2, end_column=7)
    big = ws.cell(row=hr, column=5, value=(f'{tasa}%' if tasa is not None else '—'))
    big.font = Font(name='Arial', bold=True, size=30)
    big.alignment = CENTER
    ws.merge_cells(start_row=r - 1, start_column=5, end_row=r, end_column=7)
    lbl = ws.cell(row=r - 1, column=5, value='ÍNDICE DE APROBACIÓN DEL LOTE (meta ≥ 95%)')
    lbl.font = Font(name='Arial', size=8.5, color=GRIS_NOTA)
    lbl.alignment = CENTER_WRAP
    _marco(ws, hr, r, 5, 7)

    # ---- IV. criterios de evaluación ----
    base = r + 2
    _titulo_seccion(ws, base, 2, 7, 'IV. CRITERIOS DE EVALUACIÓN Y LEYENDA')
    criterios = [
        ('Fase 1', 'La desviación estándar de las 20 mediciones del canal debe estar dentro del rango [mín, máx] del perfil. Si falla, la pieza se detiene aquí.'),
        ('Fase 2', 'Si pasó la fase 1, la media del canal debe estar dentro de su rango [mín, máx] del perfil.'),
        ('PASÓ', 'Todos los canales dentro de límites en ambas fases.'),
        ('NO PASÓ', 'Al menos un canal fuera de límites; repetir la prueba en el siguiente desechable.'),
        ('SIN ESTÁNDAR', 'Ningún canal falló, pero hay canales sin límite definido: no puede afirmarse que pasó.'),
        ('RECHAZADO', 'El archivo no tiene el formato del equipo (10 canales ópticos + 4 eléctricos); no se evaluó.'),
    ]
    r = base + 2
    for termino, texto in criterios:
        ws.row_dimensions[r].height = 22
        c1 = ws.cell(row=r, column=2, value=termino)
        c1.alignment = CENTER
        c1.border = BORDER
        c1.font = Font(name='Arial', bold=True, size=9)
        c1.fill = FILL_ETIQUETA
        _merge(ws, r, 3, 7, texto, font=Font(name='Arial', size=9), align=IZQ_WRAP)
        r += 1

    # ---- V. aprobaciones (tabla estilo registro de lote) ----
    base = r + 1
    _titulo_seccion(ws, base, 2, 7, 'V. APROBACIONES')
    hr2 = base + 2
    _encabezado_tabla(ws, hr2, [(2, 3, 'Función'), (4, 4, 'Nombre'), (5, 5, 'Firma'), (6, 7, 'Fecha')])
    firmas = [('Elaboró — Operador de Control de Calidad', batch.get('operador', '')),
              ('Revisó — Supervisor de Calidad', ''),
              ('Aprobó — Aseguramiento de Calidad', '')]
    r = hr2 + 1
    for funcion, nombre in firmas:
        ws.row_dimensions[r].height = 28
        _merge(ws, r, 2, 3, funcion, font=Font(name='Arial', bold=True, size=9), align=IZQ)
        cn = ws.cell(row=r, column=4, value=nombre)
        cn.font = NORM; cn.alignment = CENTER; cn.border = BORDER
        ws.cell(row=r, column=5).border = BORDER
        _merge(ws, r, 6, 7, None)
        r += 1
    _merge(ws, r, 2, 7, 'La liberación del lote requiere el nombre, la firma y la fecha de las tres funciones.',
           font=Font(name='Arial', size=8, color=GRIS_NOTA), align=IZQ, borde=False)

    # ---- pie ----
    pie = r + 2
    _merge(ws, pie, 2, 7,
           'Documento generado electrónicamente por el Sistema de Control de Calidad. '
           'El detalle por pieza está en la hoja "Resultados por pieza" y los datos crudos '
           'por canal en "Verificación de cálculos".',
           font=MINI, align=IZQ_WRAP, borde=False)
    ws.row_dimensions[pie].height = 22


# =================== exportación principal ===================
def exportar_lote_excel(ruta_salida, batch, resultados):
    wb = Workbook()

    total = len(resultados)
    pasaron = sum(1 for x in resultados if x['veredicto'] == 'PASO')
    incompletos = sum(1 for x in resultados if x['veredicto'] == 'INCOMPLETO')
    rechazados = sum(1 for x in resultados if x['veredicto'] == 'RECHAZADO')
    fallaron = total - pasaron - incompletos - rechazados
    fecha_lote = (batch.get('fecha', '') or '').replace('-', '')[:8]
    codigo = f"QC-LOT-{batch.get('id', '') or fecha_lote or 'S/N'}"

    # ---------- hoja 1: reporte de liberación ----------
    ws = wb.active
    ws.title = 'Reporte de liberación'
    _hoja_portada(ws, batch, (total, pasaron, fallaron, incompletos, rechazados), codigo)
    _preparar_impresion(ws, 'portrait', 8)

    # ---------- hoja 2: resultados por pieza ----------
    wr = wb.create_sheet('Resultados por pieza')
    _membrete(wr, 6, 'RESULTADOS POR PIEZA', batch, codigo)
    hr = 5
    _encabezado_tabla(wr, hr, [(i, i, h) for i, h in enumerate(
        ['No. pieza', 'Archivo', 'Óptico', 'Eléctrico', 'Veredicto final', 'Alerta'], start=1)])
    wr.freeze_panes = wr.cell(row=hr + 1, column=1)

    r = hr + 1
    ordenados = sorted(resultados, key=lambda z: str(z['pieza']))
    for veredicto in ORDEN_GRUPOS:
        grupo = [x for x in ordenados if x['veredicto'] == veredicto]
        if not grupo:
            continue
        _grupo(wr, r, 1, 6, f'{ROTULOS[veredicto]} ({len(grupo)})')
        r += 1
        for x in grupo:
            _caja(wr, r, 1, 6)
            wr.cell(row=r, column=1, value=x['pieza']).font = BOLD
            wr.cell(row=r, column=1).alignment = CENTER
            wr.cell(row=r, column=2, value=x['archivo']).font = NORM
            eo = _estado_prueba(x['detalle'], 'Optico')
            ee = _estado_prueba(x['detalle'], 'Electrico')
            _estado(wr.cell(row=r, column=3, value=eo), eo)
            _estado(wr.cell(row=r, column=4, value=ee), ee)
            if veredicto == 'PASO':
                vtxt, alerta = 'PASÓ', ''
            elif veredicto == 'INCOMPLETO':
                vtxt, alerta = 'Sin estándar', 'Completar perfil'
            elif veredicto == 'RECHAZADO':
                vtxt, alerta = 'RECHAZADO', 'Archivo inválido — reexportar el JSON del equipo'
            else:
                vtxt, alerta = 'NO PASÓ', 'Repetir prueba'
            _estado(wr.cell(row=r, column=5, value=vtxt), vtxt)
            ca = wr.cell(row=r, column=6, value=alerta)
            ca.font = NORM
            r += 1
    for col, w in zip('ABCDEF', [14, 32, 14, 14, 16, 42]):
        wr.column_dimensions[col].width = w
    _preparar_impresion(wr, 'landscape', 6)

    # ---------- hoja 3: verificación de cálculos ----------
    wd = wb.create_sheet('Verificación de cálculos')
    _membrete(wd, 9, 'VERIFICACIÓN DE CÁLCULOS', batch, codigo)
    _merge(wd, 4, 1, 9,
           'Respaldo de los cálculos hechos por la app para cada canal de cada pieza: '
           'media, desviación estándar y rango de las mediciones.',
           font=Font(name='Arial', size=9, color=GRIS_NOTA), align=IZQ, borde=False)

    fila_head = 6
    _encabezado_tabla(wd, fila_head, [(i, i, h) for i, h in enumerate(
        ['No. pieza', 'Prueba', 'Canal', 'Mediciones', 'Media',
         'Desv. estándar', 'Rango', 'Estado', 'Observación'], start=1)])
    wd.freeze_panes = wd.cell(row=fila_head + 1, column=1)

    r = fila_head + 1
    for x in sorted(resultados, key=lambda z: str(z['pieza'])):
        _grupo(wd, r, 1, 9, f"Pieza {x['pieza']} — {VER_TXT.get(x['veredicto'], x['veredicto'])}")
        r += 1
        # pieza rechazada: no hay canales evaluados; dejar constancia en el respaldo
        if x['veredicto'] == 'RECHAZADO':
            _caja(wd, r, 1, 9)
            wd.cell(row=r, column=1, value=x['pieza']).alignment = CENTER
            _estado(wd.cell(row=r, column=8, value='Rechazado'), 'Rechazado')
            wd.cell(row=r, column=9, value='Archivo rechazado — formato inválido, no se evaluó.').font = NORM
            r += 1
            continue
        for prueba, filas in x['detalle'].items():
            for f in filas:
                _caja(wd, r, 1, 9)
                wd.cell(row=r, column=1, value=x['pieza']).alignment = CENTER
                wd.cell(row=r, column=2, value='Óptica' if prueba == 'Optico' else 'Eléctrica').font = NORM
                wd.cell(row=r, column=3, value=f['canal']).font = NORM
                wd.cell(row=r, column=4, value=f.get('n', 20)).alignment = CENTER
                wd.cell(row=r, column=5, value=round(f['media'], 2)).font = NORM
                wd.cell(row=r, column=6, value=round(f['desv'], 2)).font = NORM
                wd.cell(row=r, column=7, value=round(f.get('rango', 0), 2)).font = NORM
                if f['resultado'] == 'PASA':
                    est = 'Pasa'
                elif f['resultado'] == 'FALLA':
                    est = 'Falla'
                else:
                    est = 'Sin estándar'
                _estado(wd.cell(row=r, column=8, value=est), est)
                if f['resultado'] == 'FALLA':
                    obs = f"Fase {f.get('fase_fallo')}: " + ', '.join(f.get('motivos', []))
                elif f['resultado'] == 'SIN ESTANDAR':
                    obs = 'Este canal no tiene límite definido en el perfil seleccionado (no se evaluó).'
                else:
                    obs = ''
                wd.cell(row=r, column=9, value=obs).font = NORM
                r += 1
    widths2 = [12, 12, 16, 12, 14, 16, 14, 14, 46]
    for i, w in enumerate(widths2, start=1):
        wd.column_dimensions[get_column_letter(i)].width = w
    _preparar_impresion(wd, 'landscape', 9)

    wb.save(ruta_salida)
    return ruta_salida
