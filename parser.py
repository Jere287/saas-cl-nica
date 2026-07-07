"""
parser.py — Lee el archivo JSON que sale directo del equipo y extrae los datos
de las dos pruebas, calculando media, desviacion estandar y rango por canal.

Estructura del JSON del equipo (clave 'exam'):
  Cabecera: MAC, SN, FW, TS (fecha/hora)
  IF  : lista de frecuencias electricas, ej. ["2000","3000","4000","5000"]
  IMP : lista de muestras electricas (una por desechable-medicion, IS=1..20).
        Cada muestra tiene 'IM' = lista de 4 strings (una por frecuencia),
        cada string = "electrodoAB,electrodoBC,electrodoCA".
        -> 4 frecuencias x 3 electrodos = 4 canales electricos (una columna por
           frecuencia, mezclando los 3 electrodos, igual que el calculo manual).
  OPT : lista de muestras opticas (OS=1..20). Cada una 'OM' = lista de 10 valores
        -> 10 canales opticos.

Tambien acepta .xlsx por compatibilidad (detecta por extension).
"""
import json
import statistics
import os

# Nombres reales de los 10 canales opticos del dispositivo (longitudes de onda
# del sensor espectral, en el mismo orden que llegan en OM[0..9]). Se usan para
# que los perfiles de limites del area de calidad coincidan con los datos.
CANALES_OPTICOS = ['415', '445', '480', '515', '555', '590', '630', '680', 'CLEAR', '>700']


def _nombre_optico(indice, total):
    """Nombre del canal optico i (0-based). Si hay exactamente 10 canales usa la
    longitud de onda real; si no, cae a 'Canal N' para no inventar nombres."""
    if total == len(CANALES_OPTICOS):
        return CANALES_OPTICOS[indice]
    return f'Canal {indice + 1}'


def _stats(valores):
    vals = [float(v) for v in valores if v not in (None, '')]
    if not vals:
        return None
    media = statistics.mean(vals)
    desv = statistics.stdev(vals) if len(vals) > 1 else 0.0
    rango = max(vals) - min(vals)
    return {'media': media, 'desv': desv, 'rango': rango, 'n': len(vals)}


def _leer_json(ruta):
    with open(ruta, 'r', encoding='utf-8') as f:
        d = json.load(f)
    ex = d.get('exam', d)
    resultado = {'archivo': ruta, 'cabecera': {}, 'pruebas': {}}
    resultado['cabecera'] = {
        'Equipment number': ex.get('SN', ''),
        'MAC': ex.get('MAC', ''),
        'Firmware': ex.get('FW', ''),
        'Test date': ex.get('TS', ''),
    }

    # ---- OPTICO ----
    opt_muestras = ex.get('OPT', [])
    if opt_muestras:
        n_canales = len(opt_muestras[0]['OM'])
        canales = []
        for ch in range(n_canales):
            valores = [m['OM'][ch] for m in opt_muestras if ch < len(m['OM'])]
            st = _stats(valores)
            if st:
                canales.append({'canal': _nombre_optico(ch, n_canales), **st})
        if canales:
            resultado['pruebas']['Optico'] = canales

    # ---- ELECTRICO ----
    # El estandar (manual) calcula UNA media/desviacion/rango por FRECUENCIA,
    # juntando los 3 electrodos (A-B, B-C, C-A) de esa frecuencia.
    # -> 4 canales electricos (2000, 3000, 4000, 5000 Hz), 60 valores cada uno
    #    (3 electrodos x 20 mediciones).
    imp_muestras = ex.get('IMP', [])
    frecuencias = ex.get('IF', [])
    if imp_muestras and frecuencias:
        canales = []
        for fi, freq in enumerate(frecuencias):
            valores = []
            for m in imp_muestras:
                fila = m.get('IM', [])
                if fi < len(fila):
                    for v in str(fila[fi]).split(','):
                        v = v.strip()
                        if v:
                            valores.append(v)
            st = _stats(valores)
            if st:
                canales.append({'canal': f'{freq}Hz', **st})
        if canales:
            resultado['pruebas']['Electrico'] = canales

    return resultado


def _leer_xlsx(ruta):
    # compatibilidad con el formato Excel anterior
    import statistics as _st
    from openpyxl import load_workbook
    OPT_BLOCK_COLS = [2, 6, 11, 15, 20, 24, 29, 33, 38, 42]
    ELE_BLOCK_COLS = [2, 7, 13, 18]
    DATA_START, N = 15, 20
    wb = load_workbook(ruta, data_only=True)
    resultado = {'archivo': ruta, 'cabecera': {}, 'pruebas': {}}
    if 'Optico' in wb.sheetnames:
        ws = wb['Optico']
        canales = []
        for i, col in enumerate(OPT_BLOCK_COLS):
            vals = [ws.cell(row=DATA_START + k, column=col + 2).value for k in range(N)]
            s = _stats(vals)
            if s:
                canales.append({'canal': _nombre_optico(i, len(OPT_BLOCK_COLS)), **s})
        if canales:
            resultado['pruebas']['Optico'] = canales
    if 'Electrico' in wb.sheetnames:
        ws = wb['Electrico']
        canales = []
        for col in ELE_BLOCK_COLS:
            # una columna de Impedance = un canal (los 3 electrodos A-B/B-C/C-A
            # se mezclan en la misma serie, igual que en _leer_json y que en el
            # calculo manual del area de calidad: una columna = una media/desv/rango)
            valores = []; freq = None; r = DATA_START
            while True:
                m = ws.cell(row=r, column=col).value
                if not isinstance(m, (int, float)):
                    break
                freq = ws.cell(row=r, column=col + 2).value
                imp = ws.cell(row=r, column=col + 3).value
                if imp is not None:
                    valores.append(imp)
                r += 1
            s = _stats(valores)
            if s:
                ft = f'{int(freq)}Hz' if isinstance(freq, (int, float)) else str(freq)
                canales.append({'canal': ft, **s})
        if canales:
            resultado['pruebas']['Electrico'] = canales
    return resultado


class FormatoInvalido(Exception):
    """El archivo no corresponde al formato que entrega el equipo."""


# Estructura fija que entrega el equipo: 10 canales opticos + 4 frecuencias electricas.
N_OPTICOS = 10
N_ELECTRICOS = 4


def validar_estructura(datos):
    """Lanza FormatoInvalido si el archivo no tiene la estructura del equipo.
    Asi un archivo que no es del formato NO se puede evaluar (nunca 'pasa')."""
    pruebas = datos.get('pruebas', {})
    problemas = []
    opt = pruebas.get('Optico')
    ele = pruebas.get('Electrico')
    if not opt:
        problemas.append('no se encontró la prueba óptica (sección OPT)')
    elif len(opt) != N_OPTICOS:
        problemas.append(f'la prueba óptica tiene {len(opt)} canales y se esperaban {N_OPTICOS}')
    if not ele:
        problemas.append('no se encontró la prueba eléctrica (sección IMP)')
    elif len(ele) != N_ELECTRICOS:
        problemas.append(f'la prueba eléctrica tiene {len(ele)} frecuencias y se esperaban {N_ELECTRICOS}')
    if problemas:
        raise FormatoInvalido('El archivo no tiene el formato del equipo: ' + '; '.join(problemas) + '.')


def _leer_crudo(ruta):
    # Detecta el tipo por contenido (mas robusto que solo la extension):
    # un JSON empieza con '{' o '['; un xlsx es un ZIP que empieza con 'PK'.
    try:
        with open(ruta, 'rb') as f:
            inicio = f.read(4).lstrip()
    except Exception as e:
        raise FormatoInvalido(f'No se pudo leer el archivo: {e}')
    try:
        if inicio[:1] in (b'{', b'['):
            return _leer_json(ruta)
        if inicio[:2] == b'PK':
            return _leer_xlsx(ruta)
        # respaldo: por extension
        if ruta.lower().endswith('.json'):
            return _leer_json(ruta)
        if ruta.lower().endswith(('.xlsx', '.xlsm')):
            return _leer_xlsx(ruta)
    except FormatoInvalido:
        raise
    except Exception as e:
        raise FormatoInvalido(f'No se pudo interpretar el archivo (¿está corrupto o no es JSON/Excel válido?): {e}')
    raise FormatoInvalido('El archivo no es un JSON ni un Excel válido.')


def leer_archivo(ruta):
    datos = _leer_crudo(ruta)
    validar_estructura(datos)
    return datos


if __name__ == '__main__':
    import sys
    r = leer_archivo(sys.argv[1])
    print('Cabecera:', r['cabecera'])
    for prueba, canales in r['pruebas'].items():
        print(f'\n=== {prueba} ({len(canales)}) ===')
        for ch in canales:
            print(f"  {ch['canal']:14s} media={ch['media']:.2f} desv={ch['desv']:.2f} rango={ch['rango']:.2f}")
